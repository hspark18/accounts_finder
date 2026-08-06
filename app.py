import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# --- 1. 페이지 설정 ---
st.set_page_config(page_title="결산 분석 도우미", page_icon="🔎", layout="wide")
st.title("🔎 울산중구의회 결산 심사 분석 대시보드 (연도별 비교)")
st.markdown("**정책지원관 맞춤형 핵심 쟁점 추출기** | 💡 2024년과 2025년 결산 데이터를 비교하여 예산 집행 행태를 바로잡아 보겠습니다.")

# --- 2. e-호조 엑셀 데이터 전처리 함수 ---
@st.cache_data
def parse_ehojo_file(file_obj, year):
    try:
        # 1. 상단 2줄의 불필요한 헤더를 무시하고 3번째 줄부터 데이터 로드
        df = pd.read_excel(file_obj, header=2)
        
        # 2. '예비비' 등 세부사업명이 없는 독립 항목이 이전 사업에 합산되는 것을 방지
        df.loc[df.iloc[:, 4].astype(str).str.contains('예비비', na=False), df.columns[2]] = '예비비'
        
        # 3. 병합된 셀(빈칸)을 위에서 아래로 채워넣기
        df.iloc[:, 2] = df.iloc[:, 2].ffill()
        
        # 4. '통계목'이 '○'로 시작하는 실제 집행내역 행만 추출
        df_detail = df[df.iloc[:, 4].astype(str).str.startswith('○')].copy()
        
        # 5. e-호조 양식 열 번호(Index)를 기반으로 데이터 강제 추출
        df_detail = df_detail.iloc[:, [2, 7, 8, 9, 10, 12]]
        df_detail.columns = ['세부사업', '예산현액', '지출액', '이월액', '불용액', '불용사유']
        
        # 6. 금액 데이터를 숫자형으로 강제 변환
        num_cols = ['예산현액', '지출액', '이월액', '불용액']
        for col in num_cols:
            df_detail[col] = pd.to_numeric(df_detail[col], errors='coerce').fillna(0)
            
        # 7. 세부사업별로 모든 통계목 금액 합산 및 불용사유 병합
        grouped = df_detail.groupby('세부사업').agg({
            '예산현액': 'sum',
            '지출액': 'sum',
            '이월액': 'sum',
            '불용액': 'sum',
            '불용사유': lambda x: ', '.join(set([str(i).strip() for i in x.dropna() if str(i).strip() not in ('nan', '', '0')]))
        }).reset_index()
        
        # 8. 집행률, 불용률 계산 (소수점 둘째 자리 적용)
        grouped['집행률(%)'] = (grouped['지출액'] / grouped['예산현액'] * 100).fillna(0).round(2)
        grouped['불용률(%)'] = (grouped['불용액'] / grouped['예산현액'] * 100).fillna(0).round(2)
        grouped['연도'] = year
        
        return grouped
        
    except Exception as e:
        st.error(f"{year}년도 파일을 전처리하는 중 오류가 발생했습니다: {e}")
        return pd.DataFrame()

# --- 3. 사이드바: 데이터 업로드 및 필터링 ---
with st.sidebar:
    st.header("📂 결산 데이터 업로드")
    file_24 = st.file_uploader("2024회계연도 세출결산내역 (.xls)", type=['xls', 'xlsx'])
    file_25 = st.file_uploader("2025회계연도 세출결산내역 (.xlsx)", type=['xls', 'xlsx'])
    
    st.markdown("---")
    st.header("🎯 타겟 쟁점 설정")
    target_exec_rate = st.slider("집행률 저조 기준 (%)", 0, 100, 70, 5)
    target_unused_rate = st.slider("불용률 과다 기준 (%)", 0, 50, 10, 1)
    
    st.markdown("---")
    analyze_btn = st.button("🚀 분석 실행", type="primary", use_container_width=True)

if analyze_btn:
    if file_24 is None or file_25 is None:
        st.warning("⚠️ 2024년도와 2025년도 파일을 모두 업로드한 후 '분석 실행' 버튼을 눌러주세요.")
    else:
        with st.spinner("데이터를 추출하고 쟁점을 분석 중입니다..."):
            df_24 = parse_ehojo_file(file_24, 2024)
            df_25 = parse_ehojo_file(file_25, 2025)
            
            if not df_24.empty and not df_25.empty:
                st.session_state['df_24'] = df_24
                st.session_state['df_25'] = df_25
                st.session_state['is_analyzed'] = True

# --- 4. 엑셀 내보내기 함수 ---
def export_issue_list(issue_df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결산심사_쟁점사업목록"
    
    title_font = Font(name='맑은 고딕', size=14, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    body_font = Font(name='맑은 고딕', size=11)
    align_center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    ws.append(["🚨 2025회계연도 결산심사 주요 쟁점 사업 (집행부 질의용)"])
    ws['A1'].font = title_font
    
    columns = ["세부사업명", "예산현액(원)", "지출액(원)", "집행률(%)", "불용률(%)", "불용사유", "검토요지"]
    ws.append(columns)
    
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.alignment = align_center
        cell.fill = header_fill
        cell.border = border_all
    
    ws.column_dimensions['A'].width = 40 
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50 
    
    for idx, row in issue_df.iterrows():
        review_note = ""
        if row['집행률(%)'] <= target_exec_rate:
            review_note += f"집행률 저조({row['집행률(%)']:.2f}%) "
        if row['불용률(%)'] >= target_unused_rate:
            review_note += f"불용과다 "
            
        ws.append([
            row['세부사업'], 
            f"{row['예산현액']:,}", 
            f"{row['지출액']:,}", 
            f"{row['집행률(%)']:.2f}", 
            f"{row['불용률(%)']:.2f}", 
            row.get('불용사유', ''), 
            review_note
        ])
        
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in r:
            cell.font = body_font
            cell.border = border_all
            cell.alignment = align_center if cell.column not in (1, 6, 7) else Alignment(horizontal='left', vertical='center')
            
    wb.save(output)
    return output.getvalue()


# --- 5. 화면 구성 (Tabs) ---
if st.session_state.get('is_analyzed', False):
    df_24 = st.session_state['df_24']
    df_25 = st.session_state['df_25']
    df_all = pd.concat([df_24, df_25], ignore_index=True)
    
    tab1, tab2, tab3 = st.tabs(["💰 과별 2025년 결산자료 총괄", "🚨 핵심 쟁점 추출", "📈 24 vs 25 연도별 비교"])

    # ==========================================
    # TAB 1: 2025 총괄 요약
    # ==========================================
    with tab1:
        st.subheader("업로드된 세출결산내역 총괄 (2025년도 기준)")
        total_budget = df_25['예산현액'].sum()
        total_exp = df_25['지출액'].sum()
        total_carry = df_25['이월액'].sum()
        total_unused = df_25['불용액'].sum()
        avg_exec_rate = (total_exp / total_budget * 100) if total_budget > 0 else 0
        
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("총 예산현액", f"{total_budget:,.0f} 원")
        c2.metric("총 지출액 (집행률)", f"{total_exp:,.0f} 원", f"{avg_exec_rate:.2f}%")
        c3.metric("이월액", f"{total_carry:,.0f} 원")
        c4.metric("불용액 (불용률)", f"{total_unused:,.0f} 원", f"{(total_unused/total_budget*100):.2f}%", delta_color="inverse")
        
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**집행률 하위 사업 TOP 10 (예산 5백만원 이상)**")
            low_projects = df_25[df_25['예산현액'] >= 5000000].nsmallest(10, '집행률(%)')
            st.dataframe(
                low_projects[['세부사업', '예산현액', '집행률(%)', '불용액']].style.format({
                    "예산현액": "{:,.0f}", 
                    "불용액": "{:,.0f}", 
                    "집행률(%)": "{:.2f}"
                }), 
                use_container_width=True
            )
            
        with col2:
            st.markdown("**예산 지출 구성비**")
            fig_pie = px.pie(
                values=[total_exp, total_carry, total_unused], 
                names=["지출액", "이월액", "불용액"],
                color_discrete_sequence=['#4C78A8', '#F58518', '#E45756'],
                hole=0.4
            )
            fig_pie.update_layout(margin=dict(t=0, b=0, l=0, r=0))
            st.plotly_chart(fig_pie, use_container_width=True)

    # ==========================================
    # TAB 2: 핵심 쟁점 추출
    # ==========================================
    with tab2:
        st.subheader("🚨 예산심사 및 결산질의 대비 문제 사업")
        st.caption(f"사이드바 필터 적용: **집행률 {target_exec_rate}% 이하** 또는 **불용률 {target_unused_rate}% 이상** 사업 목록")
        
        issue_df = df_25[(df_25['집행률(%)'] <= target_exec_rate) | (df_25['불용률(%)'] >= target_unused_rate)]
        
        if not issue_df.empty:
            styled_issue_df = issue_df[['세부사업', '예산현액', '지출액', '이월액', '불용액', '집행률(%)', '불용률(%)', '불용사유']].style\
                .format({
                    "예산현액": "{:,.0f}", "지출액": "{:,.0f}", "이월액": "{:,.0f}", "불용액": "{:,.0f}", 
                    "집행률(%)": "{:.2f}", "불용률(%)": "{:.2f}"
                })\
                .background_gradient(subset=['집행률(%)'], cmap='Reds_r', vmin=0, vmax=100)\
                .background_gradient(subset=['불용률(%)'], cmap='Reds', vmin=0, vmax=100)
                
            st.dataframe(styled_issue_df, use_container_width=True, height=350)
            
            st.download_button(
                label="📥 쟁점 사업 리스트 엑셀 다운로드 (보고용)",
                data=export_issue_list(issue_df),
                file_name=f"결산심사_쟁점사업_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
            # 자동 생성 질의 요지
            st.markdown("### 💡 핵심 쟁점 질의 요지 (자동 생성)")
            sample_issue = issue_df.sort_values('예산현액', ascending=False).iloc[0]
            
            st.info(f"""
**# 수치상 문제점**
* **대상사업:** [{sample_issue['세부사업']}]
* **예산 및 집행:** 예산현액 {sample_issue['예산현액']:,}원 중 지출액은 {sample_issue['지출액']:,}원으로 **집행률이 {sample_issue['집행률(%)']:.2f}%에 불과**함.
* **과다 불용발생:** 미집행으로 인한 불용액이 **{sample_issue['불용액']:,}원(불용률 {sample_issue['불용률(%)']:.2f}%)** 발생. 
* **사유 분석:** 결산서상 불용 사유는 '{sample_issue.get('불용사유', '기재 안됨')}'으로 명시되어 있으나, 당초 예산 편성 시 과다 계상되었거나 사업 추진 검토가 부실했음을 방증함.

**# 권고 및 개선안**
* 사업 계획의 타당성 원점 재검토 및 예산 집행 부진 사유에 대한 명확한 소명 요구.
* 관행적·연례 반복적인 집행 부진 사업일 경우, 내년도 본예산 편성 시 삭감 등 강력한 페널티 적용 필요.
* 예산의 사장(死藏)을 막기 위해, 집행 불가 시 추경을 통한 과감한 감액 및 삭감 편성 지적.
            """)
        else:
            st.success("지정된 기준을 초과하는 문제 사업이 없습니다.")

    # ==========================================
    # TAB 3: 연도별 비교
    # ==========================================
    with tab3:
        # st.subheader 동적 값 연동 (target_exec_rate)
        st.subheader(f"📈 2년 연속 집행 부진 사업 (집행률 {target_exec_rate}% 이하)")
        
        df_pivot = df_all.pivot_table(index='세부사업', columns='연도', values=['예산현액', '지출액', '집행률(%)']).reset_index()
        df_pivot.columns = [f"{col[0]}_{col[1]}" if col[1] else col[0] for col in df_pivot.columns]
        
        if '예산현액_2024' in df_pivot.columns and '예산현액_2025' in df_pivot.columns:
            chronic_issues = df_pivot[
                (df_pivot['예산현액_2024'] >= 10000000) & (df_pivot['예산현액_2025'] >= 10000000) &
                (df_pivot['집행률(%)_2024'] <= target_exec_rate) & (df_pivot['집행률(%)_2025'] <= target_exec_rate)
            ].sort_values('예산현액_2025', ascending=False)
            
            if not chronic_issues.empty:
                st.error(f"🚨 **2년 연속 집행률 {target_exec_rate}% 이하인 고질적 부진 사업 목록입니다. 관행적 예산 편성이 강하게 의심됩니다.**")
                styled_chronic = chronic_issues[['세부사업', '예산현액_2024', '집행률(%)_2024', '예산현액_2025', '집행률(%)_2025']].style\
                    .format({
                        "예산현액_2024": "{:,.0f}", "예산현액_2025": "{:,.0f}",
                        "집행률(%)_2024": "{:.2f}%", "집행률(%)_2025": "{:.2f}%"
                    })\
                    .background_gradient(subset=['집행률(%)_2024', '집행률(%)_2025'], cmap='Oranges_r', vmin=0, vmax=100)
                st.dataframe(styled_chronic, use_container_width=True)
            else:
                st.success(f"2년 연속 집행률 {target_exec_rate}% 이하인 고질적인 부진 사업은 발견되지 않았습니다.")
            
            st.markdown("---")
            st.markdown("**특정 세부사업 예산/집행 추이**")
            project_list = sorted(df_all['세부사업'].dropna().astype(str).unique().tolist())
            selected_proj = st.selectbox("추적할 세부사업명을 선택하세요", project_list)
            
            proj_df = df_all[df_all['세부사업'] == selected_proj]
            
            fig = go.Figure()
            fig.add_trace(go.Bar(x=proj_df['연도'].astype(str), y=proj_df['예산현액'], name='예산현액', marker_color='#93a1a1'))
            fig.add_trace(go.Bar(x=proj_df['연도'].astype(str), y=proj_df['지출액'], name='지출액', marker_color='#4C78A8'))
            fig.add_trace(go.Bar(x=proj_df['연도'].astype(str), y=proj_df['불용액'], name='불용액', marker_color='#E45756'))
            fig.update_layout(barmode='group', yaxis_title="금액 (원)", xaxis_title="회계연도")
            st.plotly_chart(fig, use_container_width=True)
            
            # 표(테이블) 형태로 연도별 수치 표시
            st.markdown("##### 📌 연도별 집행 상세 내역")
            # 선택한 사업의 데이터만 추출 후 가독성 좋게 컬럼 정리
            detail_table = proj_df[['연도', '예산현액', '지출액', '불용액', '집행률(%)', '불용률(%)']].copy()
            detail_table = detail_table.sort_values('연도', ascending=True).set_index('연도')
            
            st.dataframe(
                detail_table.style.format({
                    "예산현액": "{:,.0f} 원", 
                    "지출액": "{:,.0f} 원", 
                    "불용액": "{:,.0f} 원", 
                    "집행률(%)": "{:.2f}%", 
                    "불용률(%)": "{:.2f}%"
                }),
                use_container_width=True
            )
else:
    st.info("👈 좌측 사이드바에서 파일을 업로드하고 **[🚀 분석 실행]** 버튼을 눌러주세요.")